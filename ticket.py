import os
import re
import pdfplumber
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox

# ----------------------------
# INTERFACE GRÁFICA
# ----------------------------

def selecionar_pasta():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    pasta = filedialog.askdirectory(title="Selecione a pasta com os PDFs")
    root.destroy()
    return pasta

def selecionar_planilha():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    arquivo = filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
    )
    root.destroy()
    return arquivo

# ----------------------------
# SELEÇÃO DE CAMINHOS
# ----------------------------

PASTA_PDFS = selecionar_pasta()
if not PASTA_PDFS:
    messagebox.showerror("Erro", "Nenhuma pasta selecionada. O programa será encerrado.")
    exit()

PLANILHA_EXCEL = selecionar_planilha()
if not PLANILHA_EXCEL:
    messagebox.showerror("Erro", "Nenhuma planilha selecionada. O programa será encerrado.")
    exit()

dados = []

for arquivo in os.listdir(PASTA_PDFS):

    if not arquivo.lower().endswith(".pdf"):
        continue

    caminho_pdf = os.path.join(PASTA_PDFS, arquivo)

    razao_social = ""
    dps_serie = ""
    nfse = ""

    # ----------------------------
    # EXTRAIR TEXTO DO PDF
    # ----------------------------

    with pdfplumber.open(caminho_pdf) as pdf:

        texto = ""

        for pagina in pdf.pages:
            conteudo = pagina.extract_text()
            if conteudo:
                texto += conteudo

        # ----------------------------
        # RAZAO SOCIAL (Tomador - segunda ocorrência)
        # A primeira é do PRESTADOR (TICKET GESTAO...)
        # A segunda é do TOMADOR (que queremos)
        # ----------------------------

        matches_razao = re.findall(r"Nome/Razão Social:\s*(.+)", texto)

        if len(matches_razao) >= 2:
            razao_social = matches_razao[1].strip()
        elif len(matches_razao) == 1:
            razao_social = matches_razao[0].strip()

        # ----------------------------
        # NFS-e e DPS / SERIE
        # Os valores estão na LINHA ABAIXO do cabeçalho
        # ----------------------------

        linhas = texto.split('\n')
        
        # DEBUG: mostrar primeiras 20 linhas
        print("="*50)
        print(f"LINHAS DO PDF {arquivo}:")
        for idx, l in enumerate(linhas[:20]):
            print(f"  [{idx}] {l}")
        print("="*50)
        
        for i, linha in enumerate(linhas):
            # NFS-e Nacional - valor pode estar algumas linhas depois
            if "Número NFS-e Nacional" in linha:
                # Procura nas próximas 3 linhas por um número grande (NFS-e tem muitos dígitos)
                for j in range(1, 4):
                    if i + j < len(linhas):
                        proxima = linhas[i + j].strip()
                        match_nfse = re.search(r"(\d{6,})", proxima)  # pelo menos 6 dígitos
                        if match_nfse:
                            nfse = match_nfse.group(1)
                            break
            
            # DPS / Série - valor pode estar algumas linhas depois (formato: 1692903 / T03)
            if "Número DPS" in linha and "Série DPS" in linha:
                for j in range(1, 6):  # procura até 5 linhas depois
                    if i + j < len(linhas):
                        proxima = linhas[i + j].strip()
                        # Captura número / código (série pode ter letras)
                        match_dps = re.search(r"(\d+)\s*/\s*([A-Za-z0-9]+)", proxima)
                        if match_dps:
                            dps_serie = f"{match_dps.group(1)} / {match_dps.group(2)}"
                            break

    # limpar caracteres inválidos para nome de arquivo
    razao_limpa = re.sub(r'[\\/*?:"<>|]', "", razao_social)

    # ----------------------------
    # RENOMEAR ARQUIVO
    # ----------------------------

    novo_nome = arquivo.replace(".pdf", f"_{razao_limpa}.pdf")
    novo_caminho = os.path.join(PASTA_PDFS, novo_nome)

    if not os.path.exists(novo_caminho):
        os.rename(caminho_pdf, novo_caminho)

    # ----------------------------
    # GUARDAR DADOS
    # ----------------------------

    print(f"PDF: {arquivo}")
    print(f"  Razão Social: {razao_social}")
    print(f"  NFS-e: {nfse}")
    print(f"  DPS/Série: {dps_serie}")
    print()

    dados.append({
        "Razao Social": razao_social,
        "Numero NFSe": nfse,
        "Numero DPS / Serie DPS": dps_serie
    })

# ----------------------------
# PREENCHER PLANILHA EXISTENTE
# ----------------------------

wb = load_workbook(PLANILHA_EXCEL)
ws = wb.active

# Encontrar primeira linha vazia a partir da linha 2 (assumindo cabeçalho na linha 1)
linha_inicial = 2
for row in range(2, ws.max_row + 2):
    if ws.cell(row=row, column=2).value is None:
        linha_inicial = row
        break

# Preencher colunas B (Razão Social), C (NFS-e Nacional) e D (DPS / Série)
for i, registro in enumerate(dados):
    linha = linha_inicial + i
    ws.cell(row=linha, column=2, value=registro["Razao Social"])           # Coluna B
    ws.cell(row=linha, column=3, value=registro["Numero NFSe"])            # Coluna C
    ws.cell(row=linha, column=4, value=registro["Numero DPS / Serie DPS"]) # Coluna D
    print(f"Linha {linha}: {registro['Razao Social'][:40]}...")

wb.save(PLANILHA_EXCEL)

messagebox.showinfo("Concluído", f"Processamento finalizado!\n{len(dados)} registro(s) processado(s).")
print(f"\nProcessamento finalizado! {len(dados)} registro(s) processado(s).")