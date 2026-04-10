import os
import re
import gc
import uuid
import time
import pdfplumber
import logging
import zipfile
from io import BytesIO
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.utils import secure_filename

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB max

# Criar pasta de uploads se não existir
try:
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs('outputs', exist_ok=True)
    logger.info("Pastas criadas com sucesso")
except Exception as e:
    logger.error(f"Erro ao criar pastas: {e}")


def _cleanup_old_files():
    """Remove arquivos temporários com mais de 1 hora."""
    try:
        now = time.time()
        for folder in [app.config['UPLOAD_FOLDER'], 'outputs']:
            if os.path.exists(folder):
                for f in os.listdir(folder):
                    fpath = os.path.join(folder, f)
                    if os.path.isfile(fpath) and (now - os.path.getmtime(fpath)) > 3600:
                        try:
                            os.remove(fpath)
                        except Exception:
                            pass
    except Exception as e:
        logger.warning(f"Erro ao limpar arquivos antigos: {e}")


_cleanup_old_files()


# =============================================================================
# FUNÇÕES DE EXTRAÇÃO - TICKET
# =============================================================================

def extrair_dados_pdf_ticket(caminho_pdf):
    """Extrai dados de um PDF de fatura Ticket."""
    razao_social = ""
    dps_serie = ""
    nfse = ""
    valor_total_nf = ""
    valor_liquido_nf = ""

    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            texto = ""
            # Processar apenas as primeiras 3 páginas para economia de memória
            max_pages = min(3, len(pdf.pages))
            for i in range(max_pages):
                page = pdf.pages[i]
                conteudo = page.extract_text()
                if conteudo:
                    texto += conteudo
                del page

            # RAZAO SOCIAL (Tomador - segunda ocorrência)
            matches_razao = re.findall(r"Nome/Razão Social:\s*(.+)", texto)
            if len(matches_razao) >= 2:
                razao_social = matches_razao[1].strip()
            elif len(matches_razao) == 1:
                razao_social = matches_razao[0].strip()

            # NFS-e e DPS / SERIE
            linhas = texto.split('\n')

            for i, linha in enumerate(linhas):
                # NFS-e Nacional
                if "Número NFS-e Nacional" in linha:
                    for j in range(1, 4):
                        if i + j < len(linhas):
                            proxima = linhas[i + j].strip()
                            match_nfse = re.search(r"(\d{6,})", proxima)
                            if match_nfse:
                                nfse = match_nfse.group(1)
                                break

                # DPS / Série
                if "Número DPS" in linha and "Série DPS" in linha:
                    for j in range(1, 6):
                        if i + j < len(linhas):
                            proxima = linhas[i + j].strip()
                            match_dps = re.search(r"(\d+)\s*/\s*([A-Za-z0-9]+)", proxima)
                            if match_dps:
                                dps_serie = f"{match_dps.group(1)} / {match_dps.group(2)}"
                                break

                # Valor Total da Nota Fiscal
                if "VALOR TOTAL DA NOTA FISCAL" in linha.upper():
                    match_valor = re.search(r'R\$\s*([\d.]+,\d{2})', linha)
                    if not match_valor:
                        match_valor = re.search(r'([\d.]+,\d{2})\s*$', linha.strip())
                    if match_valor:
                        valor_total_nf = match_valor.group(1)

                # Valor Líquido da Nota Fiscal
                if "VALOR L\u00cdQUIDO DA NOTA FISCAL" in linha.upper() or "VALOR LIQUIDO DA NOTA FISCAL" in linha.upper():
                    match_valor = re.search(r'R\$\s*([\d.]+,\d{2})', linha)
                    if not match_valor:
                        match_valor = re.search(r'([\d.]+,\d{2})\s*$', linha.strip())
                    if match_valor:
                        valor_liquido_nf = match_valor.group(1)
    except Exception as e:
        logger.error(f"Erro ao extrair dados do PDF: {e}")
        raise

    return {
        "razao_social": razao_social,
        "nfse": nfse,
        "dps_serie": dps_serie,
        "valor_total_nf": valor_total_nf,
        "valor_liquido_nf": valor_liquido_nf
    }


# =============================================================================
# FUNÇÕES DE EXTRAÇÃO - SEM PARAR
# =============================================================================

def normalizar_cnpj(cnpj):
    """Remove máscara do CNPJ, mantendo apenas os dígitos."""
    if not cnpj:
        return ""
    return re.sub(r'\D', '', str(cnpj))


def extrair_dados_pdf_semparar(caminho_pdf):
    """Extrai dados de um PDF de fatura Sem Parar."""
    cnpj = ""
    cnpj_normalizado = ""
    numero_fatura = ""
    numero_nota_fiscal = ""
    razao_social = ""
    valor_liquido_pagar = ""

    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            texto = ""
            num_pages = len(pdf.pages)

            # Processar as primeiras 2 páginas para cabeçalho/dados gerais
            max_pages = min(2, num_pages)
            paginas_lidas = set()
            for i in range(max_pages):
                page = pdf.pages[i]
                conteudo = page.extract_text()
                if conteudo:
                    texto += conteudo + "\n"
                paginas_lidas.add(i)
                del page

            # Garantir que a última página também é lida (Valor Líquido a Pagar)
            last_idx = num_pages - 1
            if last_idx not in paginas_lidas:
                last_page = pdf.pages[last_idx]
                conteudo_last = last_page.extract_text()
                if conteudo_last:
                    texto += conteudo_last + "\n"
                del last_page

            if not texto.strip():
                return {
                    "cnpj": "",
                    "cnpj_normalizado": "",
                    "numero_fatura": "",
                    "numero_nota_fiscal": "",
                    "razao_social": "",
                    "valor_liquido_pagar": ""
                }

            # Extração do CNPJ
            match_cnpj = re.search(r'CNPJ:\s*([\d\.\-/]+)', texto)
            if match_cnpj:
                cnpj = match_cnpj.group(1).strip()
                cnpj_normalizado = normalizar_cnpj(cnpj)

            # Extração do Número da Fatura
            match_fatura = re.search(r'N[ºo°]\s*da\s*Fatura:\s*(\d+)', texto, re.IGNORECASE)
            if match_fatura:
                numero_fatura = match_fatura.group(1).strip()

            # Extração do Número da Nota Fiscal
            match_nf = re.search(r'N[ºo°]\s*da\s*Nota\s*Fiscal:\s*(\d+)', texto, re.IGNORECASE)
            if match_nf:
                numero_nota_fiscal = match_nf.group(1).strip()

            # Extração da Razão Social (Nome)
            match_nome = re.search(r'Nome:\s*(.+?)(?:\n|$)', texto)
            if match_nome:
                razao_social = match_nome.group(1).strip()

            # Extração do Valor Líquido a Pagar (geralmente na última página)
            linhas_sp = texto.split('\n')
            for i, linha in enumerate(linhas_sp):
                linha_upper = linha.upper()
                # Aceita variações com/sem acento, com/sem dois-pontos, maiúsculas/minúsculas
                if ('LIQUIDO' in linha_upper or 'L\u00cdQUIDO' in linha_upper) and 'PAGAR' in linha_upper:
                    logger.debug(f"[SemParar] Linha candidata valor: {repr(linha)}")
                    # Tenta na mesma linha e nas próximas 3 linhas
                    for offset in range(0, 4):
                        if i + offset >= len(linhas_sp):
                            break
                        candidato = linhas_sp[i + offset]
                        logger.debug(f"[SemParar] Candidato [{offset}]: {repr(candidato)}")
                        match_valor = re.search(r'R\$\s*([\d.]+,\d{2})', candidato)
                        if not match_valor:
                            match_valor = re.search(r'\b(\d{1,3}(?:\.\d{3})*,\d{2})\b', candidato)
                        if match_valor:
                            valor_liquido_pagar = match_valor.group(1)
                            break
                    if valor_liquido_pagar:
                        break
    except Exception as e:
        logger.error(f"Erro ao extrair dados do PDF Sem Parar: {e}")
        raise

    return {
        "cnpj": cnpj,
        "cnpj_normalizado": cnpj_normalizado,
        "numero_fatura": numero_fatura,
        "numero_nota_fiscal": numero_nota_fiscal,
        "razao_social": razao_social,
        "valor_liquido_pagar": valor_liquido_pagar
    }


# =============================================================================
# ROTAS
# =============================================================================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/health')
def health():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'uploads_folder': os.path.exists(app.config['UPLOAD_FOLDER']),
        'outputs_folder': os.path.exists('outputs')
    })


@app.route('/processar', methods=['POST'])
def processar():
    """Processa PDFs do Ticket."""
    if 'arquivos' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    arquivos = request.files.getlist('arquivos')
    
    if not arquivos or arquivos[0].filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400

    dados = []
    arquivos_renomeados = []
    total = len(arquivos)

    for idx, arquivo in enumerate(arquivos):
        if arquivo and arquivo.filename.lower().endswith('.pdf'):
            # Pegar apenas o nome do arquivo, sem o caminho da pasta
            filename_original = os.path.basename(arquivo.filename)
            filename = secure_filename(filename_original)
            caminho = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            arquivo.save(caminho)

            try:
                resultado = extrair_dados_pdf_ticket(caminho)
                
                # Renomear arquivo com razão social
                if resultado.get('razao_social') and resultado['razao_social'] != '':
                    razao_social = sanitizar_nome_arquivo(resultado['razao_social'])
                    nome_base, extensao = os.path.splitext(filename)
                    
                    # Verificar se já tem a razão social no nome
                    if razao_social.lower() not in nome_base.lower():
                        novo_nome = f"{nome_base}_{razao_social}{extensao}"
                        novo_caminho = os.path.join(app.config['UPLOAD_FOLDER'], novo_nome)
                        
                        # Evitar sobrescrever arquivo existente
                        contador = 1
                        while os.path.exists(novo_caminho):
                            novo_nome = f"{nome_base}_{razao_social}_{contador}{extensao}"
                            novo_caminho = os.path.join(app.config['UPLOAD_FOLDER'], novo_nome)
                            contador += 1
                        
                        os.rename(caminho, novo_caminho)
                        arquivos_renomeados.append({
                            'original': filename,
                            'novo': novo_nome
                        })
                        resultado['arquivo'] = novo_nome
                        resultado['arquivo_renomeado'] = True
                    else:
                        resultado['arquivo'] = filename
                        resultado['arquivo_renomeado'] = False
                else:
                    resultado['arquivo'] = filename
                    resultado['arquivo_renomeado'] = False
                    
                dados.append(resultado)
            except Exception as e:
                logger.error(f"Erro ao processar {filename}: {str(e)}")
                dados.append({
                    'arquivo': filename,
                    'razao_social': f'Erro: {str(e)}',
                    'nfse': '',
                    'dps_serie': '',
                    'valor_total_nf': '',
                    'valor_liquido_nf': '',
                    'arquivo_renomeado': False
                })
            finally:
                # Remover arquivo temporário após extração para liberar espaço
                for f_to_del in [caminho]:
                    if os.path.exists(f_to_del) and not any(
                        d.get('arquivo') not in (None, filename) for d in dados
                    ):
                        try:
                            os.remove(f_to_del)
                        except Exception:
                            pass
                gc.collect()

    return jsonify({
        'success': True,
        'dados': dados,
        'total': len(dados),
        'arquivos_renomeados': arquivos_renomeados
    })

# Rota desabilitada: não funciona quando deployado na nuvem
# O navegador não pode acessar sistema de arquivos local do usuário
'''
@app.route('/processar-pasta-ticket', methods=['POST'])
def processar_pasta_ticket():
    """Processa PDFs do Ticket de uma pasta local e renomeia os arquivos."""
    try:
        data = request.get_json()
        caminho_pasta = data.get('caminho_pasta', '').strip()
        renomear = data.get('renomear', False)

        if not caminho_pasta:
            return jsonify({'error': 'Caminho da pasta não informado'}), 400

        # Converter barras para o sistema operacional
        caminho_pasta = caminho_pasta.replace('/', os.sep).replace('\\', os.sep)

        if not os.path.exists(caminho_pasta):
            return jsonify({'error': f'Pasta não encontrada: {caminho_pasta}'}), 400

        if not os.path.isdir(caminho_pasta):
            return jsonify({'error': f'O caminho não é uma pasta: {caminho_pasta}'}), 400

        # Listar PDFs na pasta
        arquivos_pdf = [f for f in os.listdir(caminho_pasta) 
                       if f.lower().endswith('.pdf')]

        if not arquivos_pdf:
            return jsonify({'error': 'Nenhum arquivo PDF encontrado na pasta'}), 400

        dados = []
        arquivos_renomeados = []

        for arquivo in arquivos_pdf:
            caminho_completo = os.path.join(caminho_pasta, arquivo)
            
            try:
                resultado = extrair_dados_pdf_ticket(caminho_completo)
                resultado['arquivo'] = arquivo
                dados.append(resultado)

                # Renomear arquivo se solicitado e se tem razão social
                if renomear and resultado.get('razao_social'):
                    razao_social = sanitizar_nome_arquivo(resultado['razao_social'])
                    
                    # Obter nome base e extensão
                    nome_base, extensao = os.path.splitext(arquivo)
                    
                    # Verificar se já tem a razão social no nome
                    if razao_social.lower() not in nome_base.lower():
                        novo_nome = f"{nome_base}_{razao_social}{extensao}"
                        novo_caminho = os.path.join(caminho_pasta, novo_nome)
                        
                        # Evitar sobrescrever arquivo existente
                        contador = 1
                        while os.path.exists(novo_caminho):
                            novo_nome = f"{nome_base}_{razao_social}_{contador}{extensao}"
                            novo_caminho = os.path.join(caminho_pasta, novo_nome)
                            contador += 1
                        
                        os.rename(caminho_completo, novo_caminho)
                        arquivos_renomeados.append({
                            'original': arquivo,
                            'novo': novo_nome
                        })
                        # Atualizar nome no resultado
                        resultado['arquivo'] = novo_nome

            except Exception as e:
                dados.append({
                    'arquivo': arquivo,
                    'razao_social': f'Erro: {str(e)}',
                    'nfse': '',
                    'dps_serie': ''
                })

        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados),
            'arquivos_renomeados': arquivos_renomeados
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500
'''


@app.route('/processar-semparar', methods=['POST'])
def processar_semparar():
    """Processa PDFs do Sem Parar."""
    try:
        if 'arquivos' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400

        arquivos = request.files.getlist('arquivos')
        
        if not arquivos or arquivos[0].filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400

        dados = []
        arquivos_renomeados = []
        
        logger.info(f"Processando {len(arquivos)} arquivo(s) Sem Parar")

        for idx, arquivo in enumerate(arquivos):
            if arquivo and arquivo.filename.lower().endswith('.pdf'):
                # Pegar apenas o nome do arquivo, sem o caminho da pasta
                filename_original = os.path.basename(arquivo.filename)
                filename = secure_filename(filename_original)
                caminho = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                
                logger.info(f"Salvando arquivo: {filename}")
                arquivo.save(caminho)

                try:
                    resultado = extrair_dados_pdf_semparar(caminho)
                    logger.info(f"Dados extraídos: {resultado}")
                    
                    # Renomear arquivo com razão social
                    if resultado.get('razao_social') and resultado['razao_social'] != '':
                        razao_social = sanitizar_nome_arquivo(resultado['razao_social'])
                        nome_base, extensao = os.path.splitext(filename)
                        
                        # Verificar se já tem a razão social no nome
                        if razao_social.lower() not in nome_base.lower():
                            novo_nome = f"{nome_base}_{razao_social}{extensao}"
                            novo_caminho = os.path.join(app.config['UPLOAD_FOLDER'], novo_nome)
                            
                            # Evitar sobrescrever arquivo existente
                            contador = 1
                            while os.path.exists(novo_caminho):
                                novo_nome = f"{nome_base}_{razao_social}_{contador}{extensao}"
                                novo_caminho = os.path.join(app.config['UPLOAD_FOLDER'], novo_nome)
                                contador += 1
                            
                            os.rename(caminho, novo_caminho)
                            arquivos_renomeados.append({
                                'original': filename,
                                'novo': novo_nome
                            })
                            resultado['arquivo'] = novo_nome
                            resultado['arquivo_renomeado'] = True
                        else:
                            resultado['arquivo'] = filename
                            resultado['arquivo_renomeado'] = False
                    else:
                        resultado['arquivo'] = filename
                        resultado['arquivo_renomeado'] = False
                        
                    dados.append(resultado)
                except Exception as e:
                    logger.error(f"Erro ao processar {filename}: {str(e)}")
                    dados.append({
                        'arquivo': filename,
                        'cnpj': f'Erro: {str(e)}',
                        'cnpj_normalizado': '',
                        'numero_fatura': '',
                        'numero_nota_fiscal': '',
                        'razao_social': '',
                        'valor_liquido_pagar': '',
                        'arquivo_renomeado': False
                    })
                finally:
                    # Remover arquivo temporário após extração para liberar espaço
                    for f_to_del in [caminho]:
                        if os.path.exists(f_to_del) and not any(
                            d.get('arquivo') not in (None, filename) for d in dados
                        ):
                            try:
                                os.remove(f_to_del)
                            except Exception:
                                pass
                    gc.collect()

        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados),
            'arquivos_renomeados': arquivos_renomeados
        })
    
    except Exception as e:
        logger.error(f"Erro geral ao processar Sem Parar: {str(e)}")
        return jsonify({'error': f'Erro ao processar: {str(e)}'}), 500


def sanitizar_nome_arquivo(nome):
    """Remove caracteres inválidos para nomes de arquivo."""
    # Remove caracteres inválidos para Windows/Linux
    nome_limpo = re.sub(r'[<>:"/\\|?*]', '', nome)
    # Remove espaços extras
    nome_limpo = re.sub(r'\s+', ' ', nome_limpo).strip()
    # Limita o tamanho
    if len(nome_limpo) > 100:
        nome_limpo = nome_limpo[:100]
    return nome_limpo


# Rota desabilitada: não funciona quando deployado na nuvem
# O navegador não pode acessar sistema de arquivos local do usuário
'''
@app.route('/processar-pasta-semparar', methods=['POST'])
def processar_pasta_semparar():
    """Processa PDFs do Sem Parar de uma pasta local e renomeia os arquivos."""
    try:
        data = request.get_json()
        caminho_pasta = data.get('caminho_pasta', '').strip()
        renomear = data.get('renomear', False)

        if not caminho_pasta:
            return jsonify({'error': 'Caminho da pasta não informado'}), 400

        # Converter barras para o sistema operacional
        caminho_pasta = caminho_pasta.replace('/', os.sep).replace('\\', os.sep)

        if not os.path.exists(caminho_pasta):
            return jsonify({'error': f'Pasta não encontrada: {caminho_pasta}'}), 400

        if not os.path.isdir(caminho_pasta):
            return jsonify({'error': f'O caminho não é uma pasta: {caminho_pasta}'}), 400

        # Listar PDFs na pasta
        arquivos_pdf = [f for f in os.listdir(caminho_pasta) 
                       if f.lower().endswith('.pdf')]

        if not arquivos_pdf:
            return jsonify({'error': 'Nenhum arquivo PDF encontrado na pasta'}), 400

        dados = []
        arquivos_renomeados = []

        for arquivo in arquivos_pdf:
            caminho_completo = os.path.join(caminho_pasta, arquivo)
            
            try:
                resultado = extrair_dados_pdf_semparar(caminho_completo)
                resultado['arquivo'] = arquivo
                dados.append(resultado)

                # Renomear arquivo se solicitado e se tem razão social
                if renomear and resultado.get('razao_social'):
                    razao_social = sanitizar_nome_arquivo(resultado['razao_social'])
                    
                    # Obter nome base e extensão
                    nome_base, extensao = os.path.splitext(arquivo)
                    
                    # Verificar se já tem a razão social no nome
                    if razao_social.lower() not in nome_base.lower():
                        novo_nome = f"{nome_base}_{razao_social}{extensao}"
                        novo_caminho = os.path.join(caminho_pasta, novo_nome)
                        
                        # Evitar sobrescrever arquivo existente
                        contador = 1
                        while os.path.exists(novo_caminho):
                            novo_nome = f"{nome_base}_{razao_social}_{contador}{extensao}"
                            novo_caminho = os.path.join(caminho_pasta, novo_nome)
                            contador += 1
                        
                        os.rename(caminho_completo, novo_caminho)
                        arquivos_renomeados.append({
                            'original': arquivo,
                            'novo': novo_nome
                        })
                        # Atualizar nome no resultado
                        resultado['arquivo'] = novo_nome

            except Exception as e:
                dados.append({
                    'arquivo': arquivo,
                    'cnpj': f'Erro: {str(e)}',
                    'cnpj_normalizado': '',
                    'numero_fatura': '',
                    'numero_nota_fiscal': '',
                    'razao_social': ''
                })

        return jsonify({
            'success': True,
            'dados': dados,
            'total': len(dados),
            'arquivos_renomeados': arquivos_renomeados
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500
'''


@app.route('/gerar-excel', methods=['POST'])
def gerar_excel():
    """Gera Excel para dados do Ticket."""
    dados = request.json.get('dados', [])
    
    if not dados:
        return jsonify({'error': 'Nenhum dado para exportar'}), 400

    # Criar planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturas Ticket"

    # Cabeçalho
    ws.cell(row=1, column=1, value="Arquivo")
    ws.cell(row=1, column=2, value="Razão Social")
    ws.cell(row=1, column=3, value="Número NFS-e Nacional")
    ws.cell(row=1, column=4, value="Número DPS / Série DPS")
    ws.cell(row=1, column=5, value="Valor Total da NF")
    ws.cell(row=1, column=6, value="Valor Líquido da NF")

    # Estilizar cabeçalho
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="6A0CAF", end_color="6A0CAF", fill_type="solid")

    # Dados
    for idx, registro in enumerate(dados, start=2):
        ws.cell(row=idx, column=1, value=registro.get('arquivo', ''))
        ws.cell(row=idx, column=2, value=registro.get('razao_social', ''))
        ws.cell(row=idx, column=3, value=registro.get('nfse', ''))
        ws.cell(row=idx, column=4, value=registro.get('dps_serie', ''))
        ws.cell(row=idx, column=5, value=registro.get('valor_total_nf', ''))
        ws.cell(row=idx, column=6, value=registro.get('valor_liquido_nf', ''))

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    # Salvar
    nome_arquivo = f"faturas_ticket_{uuid.uuid4().hex[:8]}.xlsx"
    caminho_excel = os.path.join('outputs', nome_arquivo)
    wb.save(caminho_excel)

    return jsonify({
        'success': True,
        'arquivo': nome_arquivo
    })


@app.route('/gerar-excel-semparar', methods=['POST'])
def gerar_excel_semparar():
    """Gera Excel para dados do Sem Parar."""
    dados = request.json.get('dados', [])
    
    if not dados:
        return jsonify({'error': 'Nenhum dado para exportar'}), 400

    # Criar planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturas Sem Parar"

    # Cabeçalho
    ws.cell(row=1, column=1, value="Arquivo")
    ws.cell(row=1, column=2, value="CNPJ")
    ws.cell(row=1, column=3, value="Razão Social")
    ws.cell(row=1, column=4, value="Nº Fatura")
    ws.cell(row=1, column=5, value="Nº Nota Fiscal")
    ws.cell(row=1, column=6, value="Valor Líquido a Pagar")

    # Estilizar cabeçalho
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="00B894", end_color="00B894", fill_type="solid")

    # Dados
    for idx, registro in enumerate(dados, start=2):
        ws.cell(row=idx, column=1, value=registro.get('arquivo', ''))
        ws.cell(row=idx, column=2, value=registro.get('cnpj', ''))
        ws.cell(row=idx, column=3, value=registro.get('razao_social', ''))
        ws.cell(row=idx, column=4, value=registro.get('numero_fatura', ''))
        ws.cell(row=idx, column=5, value=registro.get('numero_nota_fiscal', ''))
        ws.cell(row=idx, column=6, value=registro.get('valor_liquido_pagar', ''))

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 22

    # Salvar
    nome_arquivo = f"faturas_semparar_{uuid.uuid4().hex[:8]}.xlsx"
    caminho_excel = os.path.join('outputs', nome_arquivo)
    wb.save(caminho_excel)

    return jsonify({
        'success': True,
        'arquivo': nome_arquivo
    })


@app.route('/download/<nome_arquivo>')
def download(nome_arquivo):
    caminho = os.path.join('outputs', nome_arquivo)
    if os.path.exists(caminho):
        return send_file(
            caminho,
            as_attachment=True,
            download_name=nome_arquivo
        )
    return jsonify({'error': 'Arquivo não encontrado'}), 404


@app.route('/download-pdfs-renomeados', methods=['POST'])
def download_pdfs_renomeados():
    """Baixa os PDFs renomeados em um arquivo ZIP."""
    try:
        data = request.json
        arquivos = data.get('arquivos', [])
        
        if not arquivos:
            return jsonify({'error': 'Nenhum arquivo para download'}), 400
        
        # Criar arquivo ZIP em memória
        memory_file = BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for arquivo in arquivos:
                caminho = os.path.join(app.config['UPLOAD_FOLDER'], arquivo)
                if os.path.exists(caminho):
                    zf.write(caminho, arquivo)
        
        memory_file.seek(0)
        
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'arquivos_renomeados_{uuid.uuid4().hex[:8]}.zip'
        )
    except Exception as e:
        logger.error(f"Erro ao criar ZIP: {e}")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'true').lower() == 'true'
    app.run(debug=debug, host='0.0.0.0', port=port)
